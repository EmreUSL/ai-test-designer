import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


def export_to_excel(test_cases, file_path="output/test_cases.xlsx"):
    if not test_cases:
        print("⚠ No test cases to export.")
        return

    os.makedirs(os.path.dirname(file_path), exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"

    headers = [
        "Title",
        "Type",
        "Priority",
        "Severity",
        "Preconditions",
        "Steps",
        "Expected Result"
    ]

    ws.append(headers)

    # Header bold
    for col in ws[1]:
        col.font = Font(bold=True)

    for index, case in enumerate(test_cases, start=1):

        tc_id = f"TC-{index:03d}"
        full_title = f"{tc_id} {case.get('title', '')}"

        steps_text = "\n".join(case.get("steps", []))

        row = [
            full_title,
            case.get("type", ""),
            case.get("priority", ""),
            case.get("severity", ""),
            case.get("preconditions", ""),
            steps_text,
            case.get("expected_result", "")
        ]

        ws.append(row)

    # Wrap text for steps column
    for row in ws.iter_rows(min_row=2):
        row[5].alignment = Alignment(wrap_text=True)

    # Auto column width
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    wb.save(file_path)

    print(f"\n✅ Excel exported successfully: {file_path}")
