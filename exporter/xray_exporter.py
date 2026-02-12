import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


def export_to_xray_excel(test_cases, file_path="output/xray_import.xlsx"):
    if not test_cases:
        print("⚠ No test cases to export.")
        return

    os.makedirs(os.path.dirname(file_path), exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Xray Import"

    headers = [
        "Issue Type",
        "Test Summary",
        "Test Type",
        "Priority",
        "Preconditions",
        "Step",
        "Expected Result"
    ]

    ws.append(headers)

    # Header bold
    for col in ws[1]:
        col.font = Font(bold=True)

    for index, case in enumerate(test_cases, start=1):

        tc_id = f"TC-{index:03d}"
        summary = f"{tc_id} {case.get('title', '')}"

        steps = case.get("steps", [])
        expected = case.get("expected_result", "")

        # Eğer birden fazla step varsa her step ayrı satır
        for step in steps:
            row = [
                "Test",              # Issue Type
                summary,
                "Manual",            # Test Type
                case.get("priority", "Medium"),
                case.get("preconditions", ""),
                step,
                expected
            ]
            ws.append(row)

    # Wrap text
    for row in ws.iter_rows(min_row=2):
        row[4].alignment = Alignment(wrap_text=True)  # Preconditions
        row[5].alignment = Alignment(wrap_text=True)  # Step
        row[6].alignment = Alignment(wrap_text=True)  # Expected

    wb.save(file_path)

    print(f"\n✅ Xray Excel exported: {file_path}")
